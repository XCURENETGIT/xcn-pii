from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path(__file__).resolve().parents[1] / "docs" / "xcn-pii_사업화_차별화_요구사항_20260804.xlsx"

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
GREEN = "70AD47"
PALE_GREEN = "E2F0D9"
YELLOW = "FFD966"
PALE_YELLOW = "FFF2CC"
RED = "C00000"
PALE_RED = "FCE4D6"
GRAY = "7F8C8D"
LIGHT_GRAY = "E7E6E6"
WHITE = "FFFFFF"

thin_gray = Side(style="thin", color="B7B7B7")
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)


def title(ws, text, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30


def section(ws, row, text, end_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(color=WHITE, bold=True, size=11)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def header_row(ws, row, headers):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row, col, text)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 34


def style_table(ws, start_row, end_row, end_col, alternating=True):
    for row in range(start_row, end_row + 1):
        for col in range(1, end_col + 1):
            cell = ws.cell(row, col)
            if alternating and row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def widths(ws, values):
    for idx, value in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(idx)].width = value


def setup_print(ws, landscape=True):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "xcn-pii 사업화 차별화 요구사항"
    ws.oddFooter.right.text = "Page &P / &N"
    ws.sheet_properties.outlinePr.summaryBelow = True


requirements = [
    {
        "no": 1,
        "category": "탐지 구조",
        "requirement": "개인정보 후보 탐색, 식별번호 구조·체크섬 검증, 문맥 의미 판정을 결합한 다단계 탐지 기능을 제공해야 한다.",
        "appeal": "정규식 단독 제품과 구분되는 정확도·오탐 제어 구조",
        "type": "필수",
        "score": 15,
        "criteria": "각 단계의 수행 여부와 최종 판정에 기여한 근거가 시험 결과에서 확인될 것",
        "method": "정상·비정상 체크섬 및 문맥 유무가 섞인 블라인드 데이터로 탐지 결과 비교",
        "evidence": "시험성적서, 탐지 결과 원본, 구성 설명서",
        "status": "구현",
        "priority": "최상",
        "caution": "특정 라이브러리명이나 내부 알고리즘 명칭을 강제하지 말고 동등 기능을 허용",
    },
    {
        "no": 2,
        "category": "판정 설명성",
        "requirement": "각 탐지 결과에 구조 유효성, 탐지 방식, 문맥 점수·판정 방법 등 최종 판정 근거를 반환해야 한다(동등 의미 항목 허용).",
        "appeal": "감사·튜닝·민원 대응 시 탐지 이유를 추적 가능",
        "type": "필수",
        "score": 10,
        "criteria": "탐지 건별로 구조 유효성 및 문맥 판정 근거가 누락 없이 조회될 것",
        "method": "API 응답 및 결과 파일에서 건별 근거 필드를 확인",
        "evidence": "API 명세, 샘플 응답, 화면 또는 로그 캡처",
        "status": "구현",
        "priority": "최상",
        "caution": "필드명 자체가 아니라 의미와 추적 가능성을 평가",
    },
    {
        "no": 3,
        "category": "다국가 지원",
        "requirement": "한국 및 베트남 주요 개인정보 식별자를 현지어(악센트 포함·미포함), 영문 및 한글 문맥과 함께 탐지해야 한다.",
        "appeal": "베트남 사업·외국인 정보 처리 환경에 특화",
        "type": "필수",
        "score": 15,
        "criteria": "한국 주민·외국인 식별정보와 베트남 CCCD, 세금, 사회보험, 여권 등 합의된 유형을 시험 데이터에서 탐지할 것",
        "method": "언어·표기 변형별 양성/음성 데이터 블라인드 시험",
        "evidence": "지원 유형표, 언어별 시험 결과, 탐지 샘플",
        "status": "구현",
        "priority": "최상",
        "caution": "최종 지원 유형과 표기 변형은 납품 범위에 맞춰 별첨으로 확정",
    },
    {
        "no": 4,
        "category": "오탐 억제",
        "requirement": "동일한 숫자열이라도 개인정보 문맥과 주문번호·정책번호·마스킹 문자열 등 비개인정보 문맥을 구분하는 문맥 기반 오탐 억제 기능을 제공해야 한다.",
        "appeal": "고객이 체감하는 오탐 감소를 직접 증명",
        "type": "필수",
        "score": 15,
        "criteria": "동일 숫자열의 양성/음성 쌍 테스트에서 사전 합의한 정탐률·오탐률 기준을 만족할 것",
        "method": "라벨만 다른 최소 100쌍 이상의 페어 데이터로 정탐률·정밀도·재현율 측정",
        "evidence": "블라인드 데이터 해시, 정답지, 혼동행렬, 산출식",
        "status": "구현",
        "priority": "최상",
        "caution": "수치 기준은 실제 고객 데이터 재시험 후 확정하고 데이터 누수를 방지",
    },
    {
        "no": 5,
        "category": "폐쇄망 운영",
        "requirement": "GPU와 외부 인터넷 연결 없이 일반 CPU 기반 폐쇄망 환경에서 모든 탐지 기능이 동작해야 한다.",
        "appeal": "공공·금융 폐쇄망 도입 비용과 보안 부담 절감",
        "type": "필수",
        "score": 10,
        "criteria": "외부 통신을 차단한 CPU 전용 환경에서 기능·성능 시험을 완료할 것",
        "method": "GPU 미장착 환경에서 외부 방화벽 차단 및 패킷 캡처와 함께 시험",
        "evidence": "시스템 사양, 네트워크 캡처, 실행 로그, 설치 가이드",
        "status": "구현",
        "priority": "최상",
        "caution": "CPU 모델·코어·메모리·입력 크기를 시험 조건에 반드시 명시",
    },
    {
        "no": 6,
        "category": "성능",
        "requirement": "지정 CPU 기준, 2,991자 입력과 동시성 10 조건에서 초당 40건 이상 및 응답시간 p95 300ms 이하를 만족해야 한다.",
        "appeal": "현재 내부 기준 성능을 정량 차별화",
        "type": "정량평가",
        "score": 15,
        "criteria": "워밍업 10건 이후 유효 요청의 처리량 ≥ 40 RPS, p95 ≤ 300ms, 오류율 0%",
        "method": "동일 장비·동일 입력·동일 도구로 3회 시험 후 중앙값 적용",
        "evidence": "부하시험 스크립트, 원시 결과, 장비 사양, 요약 보고서",
        "status": "기준 보유",
        "priority": "상",
        "caution": "내부 측정값 42.38 RPS·p95 270.13ms에 여유가 작으므로 제안 전 대상 장비에서 재검증",
    },
    {
        "no": 7,
        "category": "과부하 보호",
        "requirement": "처리 용량 초과 시 무제한 대기나 비정상 종료 대신 표준 오류, 재시도 가능 시간 및 원인 식별 정보를 반환해야 한다.",
        "appeal": "운영 장애의 연쇄 확산을 막는 백프레셔 제공",
        "type": "필수",
        "score": 5,
        "criteria": "큐 포화 시 합의된 표준 오류와 재시도 지침이 반환되고 정상 부하 복귀 후 자동 회복할 것",
        "method": "처리 한도를 초과하는 단계 부하를 인가하고 오류·회복 동작 확인",
        "evidence": "API 명세, 부하시험 로그, 모니터링 화면",
        "status": "구현",
        "priority": "상",
        "caution": "특정 코드 문자열 강제보다 상호운용 가능한 표준 상태와 의미를 요구",
    },
    {
        "no": 8,
        "category": "감사 추적",
        "requirement": "모든 탐지 응답에 적용된 룰셋의 식별자, 버전 및 갱신 시각을 포함하여 결과 재현과 변경 이력 추적이 가능해야 한다.",
        "appeal": "감사와 장애 분석 시 어떤 규칙이 적용됐는지 즉시 확인",
        "type": "필수",
        "score": 5,
        "criteria": "동일 요청의 결과에서 적용 룰 버전을 확인하고 변경 전후 이력을 재현할 수 있을 것",
        "method": "룰 갱신 전후 동일 데이터 호출 및 메타데이터·이력 비교",
        "evidence": "API 응답, 룰 변경 이력, 운영 절차서",
        "status": "구현",
        "priority": "상",
        "caution": "필드명은 동등 의미 항목 허용",
    },
    {
        "no": 9,
        "category": "무중단 운영",
        "requirement": "서비스 재기동 없이 탐지 예외 항목을 갱신하고 이후 요청부터 적용할 수 있어야 한다.",
        "appeal": "운영 중 긴급 오탐 대응 시간을 단축",
        "type": "필수",
        "score": 5,
        "criteria": "예외 갱신 중 기존 요청이 실패하지 않고 갱신 후 신규 요청에 변경 내용이 반영될 것",
        "method": "연속 트래픽 중 예외 목록 추가·교체·원복 시험",
        "evidence": "관리 API 명세, 연속 호출 로그, 변경 전후 결과",
        "status": "구현",
        "priority": "상",
        "caution": "권한 통제·감사 로그·원복 절차를 계약 범위에 포함",
    },
    {
        "no": 10,
        "category": "연계 유연성",
        "requirement": "HTTP 및 gRPC 연계 방식을 지원하고 동일 입력에 대해 탐지 의미와 결과 구조의 일관성을 보장해야 한다.",
        "appeal": "레거시와 고성능 시스템을 동시에 수용",
        "type": "선택/가점",
        "score": 5,
        "criteria": "동일 입력의 핵심 탐지 결과와 룰 버전이 프로토콜 간 일치할 것",
        "method": "동일 샘플을 두 인터페이스로 호출하여 결과 정규화 후 비교",
        "evidence": "인터페이스 명세, 비교 결과, 연계 예제",
        "status": "구현",
        "priority": "중",
        "caution": "고객 연계 구조에 따라 필수 또는 가점으로 조정",
    },
]


tests = [
    ("T-01", 1, "다단계 탐지", "정상/비정상 체크섬, 문맥 유/무 데이터 각 50건 이상", "후보 탐색→구조 검증→문맥 판정 결과를 건별 확인", "합의된 기대 결과와 일치하고 판정 근거가 확인됨", "결과 JSON/CSV, 시험 로그"),
    ("T-02", 2, "판정 설명성", "양성·음성·경계 사례 각 20건 이상", "모든 결과의 유효성·탐지 방식·문맥 근거 존재 여부 검사", "필수 의미 항목 누락 0건", "API 응답 원본, 스키마"),
    ("T-03", 3, "한국/베트남 다국어", "유형별 현지어/영문/한글 표기 변형", "표기·언어·구분자 변형별 블라인드 호출", "계약 별첨의 유형별 기준 충족", "정답지, 유형별 통계"),
    ("T-04", 4, "동일 숫자열 페어", "개인정보/비개인정보 문맥 최소 100쌍", "동일 숫자열을 문맥만 바꿔 탐지 결과 비교", "합의된 정밀도·재현율·오탐률 충족", "혼동행렬, 계산식"),
    ("T-05", 5, "CPU 폐쇄망", "GPU 미장착, 외부 통신 차단 서버", "네트워크 차단 후 설치·기동·탐지, 패킷 캡처 확인", "기능 정상, 외부 송신 0건", "사양서, pcap, 로그"),
    ("T-06", 6, "성능", "2,991자 고정 입력, 동시성 10, 워밍업 10건", "3회 부하시험 후 처리량 및 p95 산출", "≥40 RPS, p95≤300ms, 오류율 0%", "스크립트, 원시 CSV, 보고서"),
    ("T-07", 7, "큐 포화/회복", "처리 용량을 단계적으로 초과하는 부하", "큐 포화까지 부하 증가 후 정상 부하로 복귀", "표준 오류·재시도 정보 반환, 서비스 자동 회복", "클라이언트/서버 로그"),
    ("T-08", 8, "룰 버전 추적", "룰 갱신 전후 동일 요청", "갱신 전후 응답 메타데이터와 결과 비교", "버전·갱신 시각 식별 및 변경 이력 재현", "응답, 변경 이력"),
    ("T-09", 9, "무중단 예외 갱신", "연속 10 RPS 트래픽, 추가할 예외값", "트래픽 중 예외 갱신 후 결과와 오류율 확인", "중단·실패 없이 신규 요청부터 반영", "연속 호출 결과, 감사 로그"),
    ("T-10", 10, "HTTP/gRPC 동등성", "동일 샘플 100건", "두 인터페이스 결과를 정규화해 비교", "핵심 탐지 결과·룰 버전 100% 일치", "비교 CSV, API 로그"),
]


risks = [
    ("R-01", "응답 내 원문 노출", "현재 matchString 등 결과에 원문 개인정보가 포함될 수 있음", "고객 보안 심사에서 치명적 지적 가능", "응답 마스킹/해시/권한별 원문 반환 정책 구현", "보완 전 제안서에 '기본 마스킹 완료'로 기재하지 않음", "최상", "미보완"),
    ("R-02", "요청 로그 개인정보", "요청 본문 또는 탐지 원문이 로그에 남을 가능성", "로그 유출·보유기간·접근권한 위험", "구조화 로그에서 원문 제거, 마스킹, 접근통제, 보유기간 설정", "운영 로그 샘플과 설정을 납품 전 점검", "최상", "미보완"),
    ("R-03", "성능 여유", "내부 기준 42.38 RPS, p95 270.13ms로 제안 기준과 여유가 작음", "현장 장비 편차로 40 RPS/p95 300ms 미달 가능", "대상 CPU에서 3회 이상 재시험하고 안전 마진 확보", "재검증 전 수치를 확정 SLA로 사용하지 않음", "상", "재검증 필요"),
    ("R-04", "정확도 수치", "고객 대표 데이터 기준 정밀도·재현율 수치가 아직 확정되지 않음", "과도한 보장 수치가 검수 분쟁으로 연결", "블라인드 평가셋과 산출식을 계약 전에 합의", "수치 대신 시험 절차만 먼저 제시 후 실증값으로 확정", "상", "재검증 필요"),
    ("R-05", "배타적 사양 오해", "제품 내부 필드명·라이브러리·구현 방식 직접 지정 시 특정업체 사양으로 보일 수 있음", "공공입찰 이의제기 또는 사양 수정", "결과 중심 요구사항, 동등 기능 허용, 블라인드 시험 적용", "내부 문서에서만 차별화 포인트를 관리", "상", "상시 관리"),
    ("R-06", "예외 갱신 권한", "무중단 예외 갱신 API가 오용되면 탐지 우회 가능", "내부자 위협·탐지 누락", "관리자 인증, 변경 승인, 감사로그, 원복 기능 적용", "보안통제 검수 항목을 별도 추가", "상", "통제 확인 필요"),
]


sources = [
    ("제품 설계", "DESIGN_SPEC.md", "개인정보 유형, 베트남 식별자, 룰 버전, 의미 모델, 백프레셔, CPU 운영"),
    ("gRPC 제품 문서", "README_GRPC.md", "의미 탐지 모델, 성능 기준, 큐 포화 오류, 룰 메타데이터"),
    ("외부 API", "docs/EXTERNAL_API.md", "예외 목록 무중단 교체, 백프레셔 응답"),
    ("적용 현황", "docs/PII_APPLIED_STATUS_REPORT_20260606.md", "응답 원문 및 요청 로그 마스킹 보완 필요"),
    ("법령", "개인정보 보호법 제29조", "https://www.law.go.kr/LSW/lsLinkCommonInfo.do?chrClsCd=010202&lsJoLnkSeq=1033215737"),
    ("고시", "개인정보의 안전성 확보조치 기준", "https://www.law.go.kr/LSW/admRulLsInfoP.do?admRulSeq=2100000265956"),
    ("조달", "조달청 계약방법", "https://www.pps.go.kr/kor/content.do?key=00735"),
    ("계약예규", "정부 입찰·계약 집행기준", "https://www.law.go.kr/LSW/admRulLsInfoP.do?admRulNm=%28%EA%B3%84%EC%95%BD%EC%98%88%EA%B7%9C%29+%EC%A0%95%EB%B6%80+%EC%9E%85%EC%B0%B0%C2%B7%EA%B3%84%EC%95%BD+%EC%A7%91%ED%96%89%EA%B8%B0%EC%A4%80&docType=JO&joNo=001000000&languageType=KO&paras=1"),
]


wb = Workbook()
ws = wb.active
ws.title = "제안요약"
title(ws, "xcn-pii 사업화 차별화 요구사항 요약", 6)
ws.merge_cells("A3:F3")
ws["A3"] = "권고 핵심 문구"
ws["A3"].fill = PatternFill("solid", fgColor=BLUE)
ws["A3"].font = Font(color=WHITE, bold=True)
ws.merge_cells("A4:F6")
ws["A4"] = (
    "완전 폐쇄망 CPU 환경에서 한국·베트남 개인정보를 구조 검증과 문맥 의미 판정으로 탐지하고, "
    "각 결과의 판정 근거와 적용 룰 버전을 반환하며, 지정 시험 조건에서 40 RPS 이상·p95 300ms 이하를 "
    "현장 블라인드 시험으로 입증할 것. 단, 동등 이상의 기능과 성능을 제공하는 방식은 허용한다."
)
ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")
ws["A4"].fill = PatternFill("solid", fgColor=PALE_GREEN)
ws["A4"].font = Font(size=12, bold=True, color="375623")
for row in ws["A4:F6"]:
    for cell in row:
        cell.border = border
ws.row_dimensions[4].height = 35
ws.row_dimensions[5].height = 35
ws.row_dimensions[6].height = 35

section(ws, 8, "권고 평가 구성", 6)
header_row(ws, 9, ["구분", "권고값", "설명", "현재 상태", "사용 시점", "비고"])
summary_rows = [
    ("총 배점", 100, "10개 차별화 항목", "초안", "RFP/제안요청서", "고객 환경에 따라 필수/가점 조정"),
    ("핵심 필수 항목", 7, "1~9번 중 연계 유연성 제외", "대부분 구현", "기술규격서", "성능·정확도는 재시험 필요"),
    ("정량 성능", "40 RPS / p95 300ms", "2,991자, 동시성 10, 워밍업 10건", "기준 보유", "BMT/현장검수", "대상 CPU 사양을 반드시 고정"),
    ("주요 차별국가", "한국·베트남", "현지어·영문·한글 문맥", "구현", "사업 제안", "지원 식별자 목록은 별첨"),
    ("최우선 보완", "응답·로그 마스킹", "원문 개인정보 노출 통제", "미보완", "제안 전", "완료 전 과장 기재 금지"),
]
for row_idx, values in enumerate(summary_rows, 10):
    for col_idx, value in enumerate(values, 1):
        ws.cell(row_idx, col_idx, value)
style_table(ws, 10, 14, 6)
for cell in ws[14]:
    cell.fill = PatternFill("solid", fgColor=PALE_RED)
    cell.font = Font(color=RED, bold=True)

section(ws, 16, "작성·사용 원칙", 6)
principles = [
    "1. 특정 제품명, 내부 라이브러리명, 고유 필드명을 강제하지 않고 결과 중심으로 작성한다.",
    "2. '동등 이상의 기능 허용' 문구와 동일 조건 블라인드 시험을 함께 둔다.",
    "3. 성능·정확도 수치는 고객 대상 장비와 대표 데이터로 재검증한 뒤 계약 수치로 확정한다.",
    "4. 현재 보완 중인 응답·로그 개인정보 마스킹은 완료 전 충족 기능으로 표현하지 않는다.",
    "5. 본 문서는 사업·기술 초안이며 최종 공고 전 조달/법무 검토를 거친다.",
]
for idx, value in enumerate(principles, 17):
    ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=6)
    ws.cell(idx, 1, value)
    ws.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(idx, 1).border = border
    ws.row_dimensions[idx].height = 24
widths(ws, [18, 22, 42, 18, 22, 38])
setup_print(ws)


ws = wb.create_sheet("요구사항_평가표")
title(ws, "차별화 요구사항 및 권고 평가표", 13)
headers = ["No.", "분류", "요구사항 문구", "차별화·어필 포인트", "권고 구분", "배점", "합격/평가 기준", "현장 검증 방법", "필수 증빙", "구현 상태", "사업 우선순위", "법무·조달 유의사항", "검토 의견"]
header_row(ws, 3, headers)
for row_idx, item in enumerate(requirements, 4):
    values = [item["no"], item["category"], item["requirement"], item["appeal"], item["type"], item["score"], item["criteria"], item["method"], item["evidence"], item["status"], item["priority"], item["caution"], ""]
    for col_idx, value in enumerate(values, 1):
        ws.cell(row_idx, col_idx, value)
    ws.row_dimensions[row_idx].height = 92
style_table(ws, 4, 13, 13)
for row in range(4, 14):
    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row, 6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 10).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 11).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(14, 1, "합계")
ws.merge_cells(start_row=14, start_column=1, end_row=14, end_column=5)
ws.cell(14, 6, "=SUM(F4:F13)")
for col in range(1, 14):
    ws.cell(14, col).fill = PatternFill("solid", fgColor=PALE_GREEN)
    ws.cell(14, col).font = Font(bold=True)
    ws.cell(14, col).border = border
ws.cell(14, 1).alignment = Alignment(horizontal="right")
ws.cell(14, 6).alignment = Alignment(horizontal="center")
ws.auto_filter.ref = "A3:M13"
ws.freeze_panes = "C4"
ws.print_title_rows = "1:3"
widths(ws, [7, 14, 45, 28, 13, 8, 38, 38, 27, 13, 13, 40, 24])
status_dv = DataValidation(type="list", formula1='"구현,기준 보유,재검증 필요,보완 중,미구현"', allow_blank=True)
priority_dv = DataValidation(type="list", formula1='"최상,상,중,하"', allow_blank=True)
ws.add_data_validation(status_dv)
ws.add_data_validation(priority_dv)
status_dv.add("J4:J100")
priority_dv.add("K4:K100")
setup_print(ws)


ws = wb.create_sheet("현장검증_시나리오")
title(ws, "현장 BMT·검수 시나리오", 8)
header_row(ws, 3, ["시험 ID", "연계 No.", "시험 항목", "준비 데이터/환경", "시험 절차", "합격 기준", "증빙자료", "실측 결과/판정"])
for row_idx, values in enumerate(tests, 4):
    for col_idx, value in enumerate(values, 1):
        ws.cell(row_idx, col_idx, value)
    ws.row_dimensions[row_idx].height = 72
style_table(ws, 4, 13, 8)
ws.auto_filter.ref = "A3:H13"
ws.freeze_panes = "D4"
ws.print_title_rows = "1:3"
widths(ws, [11, 10, 22, 38, 42, 38, 28, 30])
setup_print(ws)


ws = wb.create_sheet("리스크_보완과제")
title(ws, "제안 전 리스크 및 제품 보완 과제", 9)
header_row(ws, 3, ["ID", "리스크", "현재 확인사항", "사업 영향", "권고 조치", "제안서 표현 원칙", "우선순위", "상태", "담당/완료예정"])
for row_idx, values in enumerate(risks, 4):
    for col_idx, value in enumerate(values, 1):
        ws.cell(row_idx, col_idx, value)
    ws.row_dimensions[row_idx].height = 76
style_table(ws, 4, 9, 9)
for row in (4, 5):
    for col in range(1, 10):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=PALE_RED)
ws.auto_filter.ref = "A3:I9"
ws.freeze_panes = "C4"
ws.print_title_rows = "1:3"
widths(ws, [10, 22, 40, 32, 42, 38, 12, 16, 22])
setup_print(ws)


ws = wb.create_sheet("근거자료")
title(ws, "내부 근거 및 공식 참고자료", 4)
header_row(ws, 3, ["구분", "자료명", "경로/URL", "활용 내용"])
for row_idx, values in enumerate(sources, 4):
    if len(values) == 3:
        category, name, usage = values
        path = name
        display_name = Path(name).name
        row_values = (category, display_name, path, usage)
    else:
        category, name, url = values
        row_values = (category, name, url, "법무·조달 검토 시 참고")
    for col_idx, value in enumerate(row_values, 1):
        ws.cell(row_idx, col_idx, value)
    if str(row_values[2]).startswith("http"):
        ws.cell(row_idx, 3).hyperlink = row_values[2]
        ws.cell(row_idx, 3).style = "Hyperlink"
    ws.row_dimensions[row_idx].height = 42
style_table(ws, 4, 11, 4)
ws.freeze_panes = "A4"
widths(ws, [16, 38, 90, 58])
setup_print(ws)


for sheet in wb.worksheets:
    sheet.sheet_properties.tabColor = BLUE

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)

# Reopen to ensure the generated OOXML package is readable and formulas/sheets survived.
check = load_workbook(OUTPUT, data_only=False)
assert check.sheetnames == ["제안요약", "요구사항_평가표", "현장검증_시나리오", "리스크_보완과제", "근거자료"]
assert check["요구사항_평가표"]["F14"].value == "=SUM(F4:F13)"
assert check["요구사항_평가표"].max_row == 14
check.close()

print(OUTPUT)
