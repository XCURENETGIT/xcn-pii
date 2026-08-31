from __future__ import annotations

import argparse
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.pii_engine.engine import detect


PII_TYPES = {
    "SN", "FN", "SSN", "DN", "PN", "MN", "BRN", "BN", "AN", "CN",
    "CPN", "CRN", "IMEI", "MCN", "EML", "VN_CCCD", "VN_MN", "VN_PN",
    "VN_TIN", "VN_SI",
}
SENSITIVE_TYPES = {
    "OTP", "API_KEY", "AUTH_TOKEN", "PASSWORD", "INTERNAL_ACCESS",
    "PRIVATE_KEY", "CLOUD_CREDENTIAL", "CONNECTION_STRING", "SIGNED_URL",
    "MFA_SECRET", "RECOVERY_CODE", "SESSION_COOKIE",
}


def _expected_positive(value: object) -> bool:
    text = str(value or "")
    return not any(marker in text for marker in ("차단되지 않는다", "0건", "비탐지된다"))


def _has_type(result: dict, allowed: set[str]) -> bool:
    return any(result.get(key) for key in allowed)


def _record(
    failures: list[tuple[str, str, bool, list[str]]],
    group: str,
    tc_id: object,
    expected: bool,
    result: dict,
    allowed: set[str],
) -> bool:
    actual = _has_type(result, allowed)
    if actual == expected:
        return True
    failures.append((group, str(tc_id), expected, sorted(key for key in allowed if result.get(key))))
    return False


def main() -> int:
    parser = argparse.ArgumentParser(description="Verify X-GEN evasion and credential XLSX cases")
    parser.add_argument("--evasion-xlsx", type=Path, required=True)
    parser.add_argument("--credential-xlsx", type=Path, required=True)
    parser.add_argument("--ruleset", default="default")
    args = parser.parse_args()

    failures: list[tuple[str, str, bool, list[str]]] = []
    summaries: list[tuple[str, int, int]] = []

    evasion = load_workbook(args.evasion_xlsx, read_only=True, data_only=False)
    passed = total = 0
    for row in evasion["서비스기준차단"].iter_rows(min_row=4, values_only=True):
        tc_id, prompt = row[2], row[8]
        if not tc_id or not prompt:
            continue
        total += 1
        result = detect(str(prompt), ruleset=args.ruleset)
        passed += _record(failures, "PII-evasion", tc_id, _expected_positive(row[9]), result, PII_TYPES)
    summaries.append(("PII-evasion", passed, total))

    credential = load_workbook(args.credential_xlsx, read_only=True, data_only=False)
    passed = total = 0
    for row in credential["엔진탐지"].iter_rows(min_row=5, values_only=True):
        tc_id, expected_type, prompt = row[2], str(row[3]), row[8]
        if not tc_id or not prompt:
            continue
        total += 1
        result = detect(str(prompt), ruleset=args.ruleset)
        expected = _expected_positive(row[9])
        actual = bool(result.get(expected_type))
        if actual == expected:
            passed += 1
        else:
            failures.append(("AUTH-engine-exact", str(tc_id), expected, sorted(key for key in SENSITIVE_TYPES if result.get(key))))
    summaries.append(("AUTH-engine-exact", passed, total))

    passed = total = 0
    for row in credential["독립검증"].iter_rows(min_row=5, values_only=True):
        tc_id, expected_type, prompt = row[2], str(row[3]), row[8]
        if not tc_id or not prompt:
            continue
        total += 1
        result = detect(str(prompt), ruleset=args.ruleset)
        passed += _record(failures, "AUTH-independent-semantic", tc_id, expected_type != "-", result, SENSITIVE_TYPES)
    summaries.append(("AUTH-independent-semantic", passed, total))

    passed = total = 0
    for row in credential["서비스기준차단"].iter_rows(min_row=4, max_row=75, values_only=True):
        tc_id, prompt = row[2], row[8]
        if not tc_id or not prompt:
            continue
        total += 1
        result = detect(str(prompt), ruleset=args.ruleset)
        passed += _record(failures, "AUTH-service", tc_id, _expected_positive(row[9]), result, SENSITIVE_TYPES)
    summaries.append(("AUTH-service", passed, total))

    for group, passed, total in summaries:
        print(f"{group}: {passed}/{total}")
    print(f"TOTAL: {sum(item[1] for item in summaries)}/{sum(item[2] for item in summaries)}")
    for failure in failures:
        print("FAIL:", failure)
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
