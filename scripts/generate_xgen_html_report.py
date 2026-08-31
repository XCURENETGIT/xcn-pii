from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
import json
from pathlib import Path
import time
from typing import Any
from urllib import request

from openpyxl import load_workbook


PII_TYPES = {
    "SN", "FN", "SSN", "DN", "PN", "MN", "BRN", "BN", "AN", "CN",
    "CPN", "CRN", "IMEI", "MCN", "EML", "VN_CCCD", "VN_MN", "VN_PN",
    "VN_TIN", "VN_SI",
}
SENSITIVE_TYPES = {
    "OTP", "API_KEY", "AUTH_TOKEN", "PASSWORD", "INTERNAL_ACCESS",
    "PRIVATE_KEY", "CLOUD_CREDENTIAL", "CONNECTION_STRING", "SIGNED_URL",
    "MFA_SECRET", "RECOVERY_CODE", "SESSION_COOKIE",
}

RRN_CONTEXT_CASES = (
    ("구분자 변형 / 마침표", "900101.1123459"),
    ("구분자 변형 / 슬래시", "900101/1123459"),
    ("구분자 변형 / 언더스코어", "900101_1123459"),
    ("구분자 변형 / 물결표", "900101~1123459"),
    ("구분자 변형 / 콜론", "900101:1123459"),
    ("구분자 변형 / 파이프", "900101|1123459"),
    ("구분자 변형 / 쉼표", "900101,1123459"),
    ("구분자 변형 / 해시", "900101#1123459"),
    ("구분자 변형 / 공백 포함 하이픈", "900101 - 1123459"),
    ("구분자 변형 / 이중 하이픈", "900101--1123459"),
    ("공백 변형 / 일반 공백", "900101 1123459"),
    ("공백 변형 / 전각공백 U+3000", "900101\u30001123459"),
    ("공백 변형 / NBSP U+00A0", "900101\u00a01123459"),
    ("공백 변형 / 3자리 분할", "900 101 112 3459"),
    ("공백 변형 / 4자리 분할·카드번호 위장", "9001 0111 2345 9"),
    ("접두·접미 경계 / 대괄호", "[900101-1123459]"),
    ("접두·접미 경계 / 인용부호·구분자 제거", '"9001011123459"'),
    ("접두·접미 경계 / id= 접두", "id=9001011123459"),
    ("접두·접미 경계 / 한국어 접미", "900101-1123459입니다"),
    ("접두·접미 경계 / RRN 접두", "RRN9001011123459"),
    ("접두·접미 경계 / 번호…번", "번호9001011123459번"),
)


def _text(value: Any) -> str:
    return "" if value is None else str(value)


def _expected_positive(value: Any) -> bool:
    text = _text(value)
    return not any(marker in text for marker in ("차단되지 않는다", "0건", "비탐지된다"))


def _get_json(url: str, timeout: float = 15.0) -> dict[str, Any]:
    with request.urlopen(url, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def _detect(api_url: str, prompt: str, ruleset: str) -> tuple[dict[str, Any], float]:
    payload = json.dumps(
        {"text": prompt, "max_results_per_type": 100},
        ensure_ascii=False,
    ).encode("utf-8")
    req = request.Request(
        api_url,
        data=payload,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "X-PII-RULESET": ruleset,
        },
        method="POST",
    )
    started = time.perf_counter()
    with request.urlopen(req, timeout=30.0) as response:
        result = json.loads(response.read().decode("utf-8"))
    return result, (time.perf_counter() - started) * 1000.0


def _detections(data: dict[str, Any]) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    for type_name in sorted(PII_TYPES | SENSITIVE_TYPES):
        items = data.get(type_name)
        if not isinstance(items, list) or not items:
            continue
        found.append(
            {
                "type": type_name,
                "count": len(items),
                "matches": [
                    {
                        "value": _text(item.get("matchString")),
                        "detected_by": _text(item.get("detected_by")),
                        "checksum_status": _text(item.get("checksum_status")),
                    }
                    for item in items
                    if isinstance(item, dict)
                ],
            }
        )
    return found


def _evaluate(case: dict[str, Any], data: dict[str, Any]) -> bool:
    mode = case["evaluation"]
    expected = bool(case["expected_detect"])
    if mode == "pii_any":
        actual = any(data.get(type_name) for type_name in PII_TYPES)
    elif mode == "sensitive_any":
        actual = any(data.get(type_name) for type_name in SENSITIVE_TYPES)
    elif mode == "exact":
        actual = bool(data.get(case["expected_type"]))
    else:
        raise ValueError(f"Unsupported evaluation mode: {mode}")
    case["actual_detect"] = actual
    return actual == expected


def _base_case(
    *,
    source_file: str,
    sheet: str,
    row_number: int,
    row: tuple[Any, ...],
    group: str,
) -> dict[str, Any]:
    return {
        "source_file": source_file,
        "sheet": sheet,
        "source_row": row_number,
        "group": group,
        "no": row[1],
        "tc_id": _text(row[2]),
        "priority": _text(row[10]) if len(row) > 10 else "",
        "source_test_day": _text(row[11]) if len(row) > 11 else "",
        "source_result_1": _text(row[12]) if len(row) > 12 else "",
        "source_result_2": _text(row[13]) if len(row) > 13 else "",
        "source_remark": _text(row[14]) if len(row) > 14 else "",
    }


def _load_cases(evasion_path: Path, credential_path: Path) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []

    evasion = load_workbook(evasion_path, read_only=True, data_only=False)
    sheet = evasion["서비스기준차단"]
    for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), 4):
        if not _text(row[2]).startswith("TC_"):
            continue
        case = _base_case(
            source_file=evasion_path.name,
            sheet=sheet.title,
            row_number=row_number,
            row=row,
            group="개인정보 변형·우회",
        )
        case.update(
            {
                "category": _text(row[6]),
                "detail": _text(row[7]),
                "prompt": _text(row[8]),
                "expected": _text(row[9]),
                "expected_detect": _expected_positive(row[9]),
                "expected_type": "PII",
                "evaluation": "pii_any",
                "execution": "AUTO",
            }
        )
        cases.append(case)

    for index, (detail, value) in enumerate(RRN_CONTEXT_CASES, 1):
        cases.append(
            {
                "source_file": "사용자 추가 요청",
                "sheet": "주민번호 문맥결합",
                "source_row": index,
                "group": "주민번호 문맥결합 추가검증",
                "no": index,
                "tc_id": f"TC_XGENAI_RRN_CTX_{index:03d}",
                "priority": "P1",
                "source_test_day": "",
                "source_result_1": "",
                "source_result_2": "",
                "source_remark": "명시적 주민번호 문맥을 결합한 추가 검증",
                "category": "주민등록번호",
                "detail": detail,
                "prompt": f"주민번호 {value} 확인해줘",
                "expected": "주민번호 문맥과 결합된 변형값이 SN 타입으로 1건 이상 탐지된다.",
                "expected_detect": True,
                "expected_type": "SN",
                "evaluation": "exact",
                "execution": "AUTO",
            }
        )

    credential = load_workbook(credential_path, read_only=True, data_only=False)

    sheet = credential["엔진탐지"]
    for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), 5):
        if not _text(row[2]).startswith("TC_"):
            continue
        expected_type = _text(row[3])
        case = _base_case(
            source_file=credential_path.name,
            sheet=sheet.title,
            row_number=row_number,
            row=row,
            group="인증정보 엔진탐지",
        )
        case.update(
            {
                "category": expected_type,
                "detail": _text(row[7]),
                "prompt": _text(row[8]),
                "expected": _text(row[9]),
                "expected_detect": _expected_positive(row[9]),
                "expected_type": expected_type,
                "evaluation": "exact",
                "execution": "AUTO",
            }
        )
        cases.append(case)

    sheet = credential["독립검증"]
    for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), 5):
        if not _text(row[2]).startswith("TC_"):
            continue
        expected_type = _text(row[3])
        case = _base_case(
            source_file=credential_path.name,
            sheet=sheet.title,
            row_number=row_number,
            row=row,
            group="인증정보 독립검증",
        )
        case.update(
            {
                "category": expected_type if expected_type != "-" else "비탐지 경계",
                "detail": _text(row[7]),
                "prompt": _text(row[8]),
                "expected": _text(row[9]),
                "expected_detect": expected_type != "-",
                "expected_type": expected_type,
                "evaluation": "sensitive_any",
                "execution": "AUTO",
            }
        )
        cases.append(case)

    sheet = credential["서비스기준차단"]
    for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), 4):
        if not _text(row[2]).startswith("TC_"):
            continue
        auto = row_number <= 75
        case = _base_case(
            source_file=credential_path.name,
            sheet=sheet.title,
            row_number=row_number,
            row=row,
            group="인증정보 서비스기준차단",
        )
        case.update(
            {
                "category": _text(row[6]),
                "detail": _text(row[7]),
                "prompt": _text(row[8]),
                "expected": _text(row[9]),
                "expected_detect": _expected_positive(row[9]),
                "expected_type": "SENSITIVE",
                "evaluation": "sensitive_any",
                "execution": "AUTO" if auto else "MANUAL",
                "manual_reason": "첨부파일·제품 차단 로그 등 PII HTTP API만으로 판정할 수 없는 시나리오" if not auto else "",
            }
        )
        cases.append(case)

    sheet = credential["제품연동"]
    for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), 5):
        if not _text(row[2]).startswith("TC_"):
            continue
        case = {
            "source_file": credential_path.name,
            "sheet": sheet.title,
            "source_row": row_number,
            "group": "인증정보 제품연동",
            "no": row[1],
            "tc_id": _text(row[2]),
            "category": " / ".join(filter(None, (_text(row[3]), _text(row[4]), _text(row[5])))),
            "detail": _text(row[6]),
            "prompt": _text(row[7]),
            "expected": _text(row[8]),
            "priority": _text(row[9]),
            "source_test_day": _text(row[10]),
            "source_result_1": _text(row[11]),
            "source_result_2": _text(row[12]),
            "source_remark": _text(row[13]),
            "expected_detect": None,
            "expected_type": "MANUAL",
            "evaluation": "manual",
            "execution": "MANUAL",
            "manual_reason": "X-GEN AI 정책·차단·로그 화면이 필요한 제품 연동 시나리오",
        }
        cases.append(case)

    return cases


def _render_html(report: dict[str, Any]) -> str:
    payload = json.dumps(report, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    return f"""<!doctype html>
<html lang=\"ko\">
<head>
  <meta charset=\"utf-8\">
  <meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">
  <title>X-GEN AI 업로드 테스트케이스 결과</title>
  <style>
    :root {{ --navy:#0b1830; --blue:#2563eb; --cyan:#06b6d4; --ink:#172033; --muted:#64748b; --line:#dbe3ef; --paper:#fff; --bg:#f3f6fb; --pass:#087a55; --pass-bg:#dcfce7; --fail:#b42318; --fail-bg:#fee2e2; --manual:#9a5b00; --manual-bg:#fff3cd; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; color:var(--ink); background:var(--bg); font:15px/1.55 \"Pretendard\",\"Noto Sans KR\",\"Malgun Gothic\",sans-serif; }}
    header {{ color:white; padding:40px 48px 72px; background:linear-gradient(125deg,var(--navy),#14376b 68%,#0e7490); }}
    header .eyebrow {{ color:#8bdcf1; font-size:12px; font-weight:800; letter-spacing:.14em; text-transform:uppercase; }}
    h1 {{ margin:7px 0 8px; font-size:36px; line-height:1.25; }}
    header p {{ margin:0; color:#d7e5f8; }}
    main {{ width:calc(100% - 80px); margin:-44px auto 50px; }}
    .cards {{ display:grid; grid-template-columns:repeat(5,minmax(180px,1fr)); gap:16px; }}
    .card {{ background:var(--paper); border:1px solid rgba(219,227,239,.9); border-radius:14px; padding:20px 22px; box-shadow:0 8px 24px rgba(20,43,78,.08); }}
    .card .label {{ color:var(--muted); font-size:12px; font-weight:700; }}
    .card .value {{ margin-top:4px; font-size:30px; font-weight:850; }}
    .panel {{ margin-top:14px; background:var(--paper); border:1px solid var(--line); border-radius:14px; box-shadow:0 5px 20px rgba(20,43,78,.05); overflow:hidden; }}
    .meta {{ display:grid; grid-template-columns:1.6fr 1fr 1fr; gap:18px; padding:18px 22px; border-bottom:1px solid var(--line); }}
    .meta strong {{ display:block; font-size:12px; color:var(--muted); }}
    .toolbar {{ display:grid; grid-template-columns:minmax(420px,2fr) repeat(3,minmax(230px,1fr)); gap:12px; padding:16px; background:#f8fafc; border-bottom:1px solid var(--line); }}
    input,select {{ width:100%; border:1px solid #cbd5e1; border-radius:9px; padding:9px 10px; background:white; color:var(--ink); font:inherit; }}
    .table-wrap {{ overflow:auto; max-height:calc(100vh - 430px); min-height:720px; }}
    table {{ width:100%; border-collapse:separate; border-spacing:0; min-width:2300px; table-layout:fixed; }}
    th {{ position:sticky; top:0; z-index:2; padding:12px 11px; color:#40516d; background:#edf2f8; border-bottom:1px solid #cad5e4; text-align:left; font-size:13px; white-space:nowrap; }}
    td {{ padding:12px 11px; vertical-align:top; border-bottom:1px solid #e8edf4; }}
    tbody tr:hover {{ background:#f8fbff; }}
    code,.mono {{ font-family:ui-monospace,SFMono-Regular,Consolas,monospace; font-size:12px; }}
    .tc {{ font-weight:800; color:#174ea6; white-space:nowrap; }}
    .prompt {{ white-space:pre-wrap; overflow-wrap:anywhere; }}
    .expected {{ white-space:pre-wrap; color:#46566f; }}
    .badge {{ display:inline-flex; align-items:center; border-radius:999px; padding:3px 8px; font-size:11px; font-weight:850; white-space:nowrap; }}
    .PASS {{ color:var(--pass); background:var(--pass-bg); }} .FAIL,.ERROR {{ color:var(--fail); background:var(--fail-bg); }} .MANUAL {{ color:var(--manual); background:var(--manual-bg); }}
    .type-chip {{ display:inline-block; margin:0 4px 4px 0; padding:2px 6px; border:1px solid #b9cdf2; border-radius:6px; color:#174ea6; background:#eef5ff; font-size:11px; font-weight:750; }}
    details {{ margin-top:5px; }} summary {{ color:#536783; cursor:pointer; font-size:12px; }}
    .details {{ padding:7px 9px; border-left:3px solid #b8c7dc; background:#f8fafc; white-space:pre-wrap; overflow-wrap:anywhere; font-size:12px; }}
    .result-line {{ white-space:nowrap; }}
    .footnote {{ padding:13px 16px; color:var(--muted); font-size:12px; border-top:1px solid var(--line); }}
    .empty {{ padding:36px; text-align:center; color:var(--muted); }}
    @media(max-width:1800px) {{ main {{ width:calc(100% - 30px); }} table {{ min-width:1420px; table-layout:auto; }} .table-wrap {{ min-height:0; max-height:70vh; }} .toolbar {{ grid-template-columns:minmax(260px,2fr) repeat(3,minmax(150px,1fr)); }} }}
    @media(max-width:900px) {{ .cards {{ grid-template-columns:repeat(2,1fr); }} .meta,.toolbar {{ grid-template-columns:1fr; }} header {{ padding-left:22px; padding-right:22px; }} }}
    @media print {{ body {{ background:white; }} header {{ padding:18px; }} main {{ width:100%; margin:0; }} .toolbar {{ display:none; }} .panel,.card {{ box-shadow:none; }} .table-wrap {{ max-height:none; overflow:visible; }} th {{ position:static; }} }}
  </style>
</head>
<body>
  <header><div class=\"eyebrow\">X-GEN AI · XCN-PII</div><h1>업로드 Excel 테스트케이스 결과</h1><p>Excel 상세 케이스와 문맥 결합 주민번호 변형 21건을 현재 배포 서비스에서 검증했습니다.</p></header>
  <main>
    <section class=\"cards\" id=\"cards\"></section>
    <section class=\"panel\">
      <div class=\"meta\" id=\"meta\"></div>
      <div class=\"toolbar\">
        <input id=\"search\" placeholder=\"TC ID, 유형, 세부항목, 프롬프트 검색\">
        <select id=\"group\"><option value=\"\">전체 시험군</option></select>
        <select id=\"status\"><option value=\"\">전체 결과</option><option>PASS</option><option>FAIL</option><option>MANUAL</option><option>ERROR</option></select>
        <select id=\"category\"><option value=\"\">전체 유형</option></select>
      </div>
      <div class=\"table-wrap\"><table><colgroup><col style=\"width:52px\"><col style=\"width:280px\"><col style=\"width:240px\"><col style=\"width:310px\"><col style=\"width:500px\"><col style=\"width:480px\"><col style=\"width:390px\"><col style=\"width:110px\"><col style=\"width:105px\"></colgroup><thead><tr><th>#</th><th>시험군 / 원본</th><th>TC ID</th><th>유형·세부항목</th><th>시험 예시</th><th>기대 결과</th><th>실제 결과</th><th>판정</th><th>시간</th></tr></thead><tbody id=\"rows\"></tbody></table><div class=\"empty\" id=\"empty\" hidden>조건에 맞는 테스트케이스가 없습니다.</div></div>
      <div class=\"footnote\"><span id=\"visibleCount\"></span> · AUTO는 52번 HTTP API 실측, MANUAL은 제품 화면·첨부파일·로그 확인이 필요한 케이스입니다.</div>
    </section>
  </main>
  <script id=\"report-data\" type=\"application/json\">{payload}</script>
  <script>
    const report=JSON.parse(document.getElementById('report-data').textContent), cases=report.cases;
    const esc=s=>String(s??'').replace(/[&<>\"']/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#39;'}}[c]));
    const counts=report.summary;
    document.getElementById('cards').innerHTML=[['전체 TC',counts.total],['자동 실행',counts.auto],['PASS',counts.pass],['FAIL / ERROR',counts.fail+counts.error],['수동 검증',counts.manual]].map(([l,v])=>`<div class=\"card\"><div class=\"label\">${{l}}</div><div class=\"value\">${{v.toLocaleString()}}</div></div>`).join('');
    document.getElementById('meta').innerHTML=`<div><strong>검증 대상</strong>${{esc(report.api_url)}} · ruleset=${{esc(report.ruleset)}}</div><div><strong>서비스 버전</strong><span class=\"mono\">${{esc(report.service_version)}}</span></div><div><strong>생성 시각</strong>${{esc(report.generated_at)}}</div>`;
    const group=document.getElementById('group'), category=document.getElementById('category');
    [...new Set(cases.map(x=>x.group))].sort().forEach(v=>group.insertAdjacentHTML('beforeend',`<option>${{esc(v)}}</option>`));
    [...new Set(cases.map(x=>x.category).filter(Boolean))].sort().forEach(v=>category.insertAdjacentHTML('beforeend',`<option>${{esc(v)}}</option>`));
    function actual(c) {{
      if(c.status==='MANUAL') return `<span class=\"badge MANUAL\">수동 검증</span><details><summary>사유</summary><div class=\"details\">${{esc(c.manual_reason)}}</div></details>`;
      if(c.status==='ERROR') return `<span class=\"badge ERROR\">실행 오류</span><details><summary>오류</summary><div class=\"details\">${{esc(c.error)}}</div></details>`;
      if(!c.detections.length) return `<span class=\"result-line\">탐지 없음</span>`;
      const chips=c.detections.map(d=>`<span class=\"type-chip\">${{esc(d.type)}} ${{d.count}}</span>`).join('');
      const lines=c.detections.flatMap(d=>d.matches.map(m=>`${{d.type}} · ${{m.value||'(값 없음)'}}${{m.detected_by?' · '+m.detected_by:''}}${{m.checksum_status?' · '+m.checksum_status:''}}`)).join('\\n');
      return chips+`<details><summary>탐지 상세</summary><div class=\"details mono\">${{esc(lines)}}</div></details>`;
    }}
    function source(c) {{ const old=[c.source_result_1,c.source_result_2,c.source_remark].filter(Boolean).join('\\n'); return `${{esc(c.group)}}<br><span class=\"mono\">${{esc(c.source_file)}} · ${{esc(c.sheet)}}!row ${{c.source_row}}</span>${{old?`<details><summary>원본 결과/비고</summary><div class=\"details\">${{esc(old)}}</div></details>`:''}}`; }}
    function render() {{
      const q=document.getElementById('search').value.trim().toLowerCase(), st=document.getElementById('status').value;
      const filtered=cases.filter(c=>(!group.value||c.group===group.value)&&(!st||c.status===st)&&(!category.value||c.category===category.value)&&(!q||[c.tc_id,c.category,c.detail,c.prompt,c.expected,c.group].join(' ').toLowerCase().includes(q)));
      document.getElementById('rows').innerHTML=filtered.map((c,i)=>`<tr><td>${{i+1}}</td><td>${{source(c)}}</td><td><span class=\"tc mono\">${{esc(c.tc_id)}}</span><br><span class=\"badge ${{c.execution==='AUTO'?'PASS':'MANUAL'}}\">${{c.execution}}</span></td><td><strong>${{esc(c.category)}}</strong><br>${{esc(c.detail)}}</td><td class=\"prompt\">${{esc(c.prompt)}}</td><td class=\"expected\">${{esc(c.expected)}}</td><td>${{actual(c)}}</td><td><span class=\"badge ${{c.status}}\">${{c.status}}</span></td><td class=\"mono\">${{c.elapsed_ms==null?'—':c.elapsed_ms.toFixed(1)+' ms'}}</td></tr>`).join('');
      document.getElementById('empty').hidden=filtered.length>0; document.getElementById('visibleCount').textContent=`표시 ${{filtered.length.toLocaleString()}} / 전체 ${{cases.length.toLocaleString()}}건`;
    }}
    ['search','group','status','category'].forEach(id=>document.getElementById(id).addEventListener(id==='search'?'input':'change',render)); render();
  </script>
</body>
</html>
"""


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a standalone HTML report for uploaded X-GEN workbooks")
    parser.add_argument("--evasion-xlsx", type=Path, required=True)
    parser.add_argument("--credential-xlsx", type=Path, required=True)
    parser.add_argument("--api-url", default="http://10.100.40.52:18005/pii/detect")
    parser.add_argument("--ruleset", default="default")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    cases = _load_cases(args.evasion_xlsx, args.credential_xlsx)
    version_url = args.api_url.rsplit("/pii/detect", 1)[0] + "/pii/version"
    version_result = _get_json(version_url)
    service_version = _text(version_result.get("app_version"))

    for index, case in enumerate(cases, 1):
        case["report_no"] = index
        case["elapsed_ms"] = None
        case["detections"] = []
        if case["execution"] == "MANUAL":
            case["status"] = "MANUAL"
            continue
        try:
            response, elapsed_ms = _detect(args.api_url, case["prompt"], args.ruleset)
            data = response.get("data") if isinstance(response, dict) else {}
            if not isinstance(data, dict):
                data = {}
            case["elapsed_ms"] = round(elapsed_ms, 3)
            case["detections"] = _detections(data)
            case["status"] = "PASS" if _evaluate(case, data) else "FAIL"
        except Exception as exc:  # Report the individual failure and continue with the remaining workbook rows.
            case["status"] = "ERROR"
            case["error"] = f"{type(exc).__name__}: {exc}"

    status_counts = Counter(case["status"] for case in cases)
    report = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "api_url": args.api_url,
        "ruleset": args.ruleset,
        "service_version": service_version,
        "sources": [
            {"name": args.evasion_xlsx.name, "path": str(args.evasion_xlsx)},
            {"name": args.credential_xlsx.name, "path": str(args.credential_xlsx)},
        ],
        "summary": {
            "total": len(cases),
            "auto": sum(case["execution"] == "AUTO" for case in cases),
            "pass": status_counts["PASS"],
            "fail": status_counts["FAIL"],
            "error": status_counts["ERROR"],
            "manual": status_counts["MANUAL"],
        },
        "group_summary": [
            {
                "group": group,
                "total": sum(case["group"] == group for case in cases),
                "pass": sum(case["group"] == group and case["status"] == "PASS" for case in cases),
                "fail": sum(case["group"] == group and case["status"] in {"FAIL", "ERROR"} for case in cases),
                "manual": sum(case["group"] == group and case["status"] == "MANUAL" for case in cases),
            }
            for group in dict.fromkeys(case["group"] for case in cases)
        ],
        "cases": cases,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(_render_html(report), encoding="utf-8")
    print(json.dumps({"output": str(args.output), **report["summary"], "version": service_version}, ensure_ascii=False))
    return 1 if status_counts["FAIL"] or status_counts["ERROR"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
